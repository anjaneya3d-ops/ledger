"""
Order Tracker — Meesho & Flipkart
Reliable order/payment tracking with verified calculations.
v2: adds authentication, API integration stubs, full record details.
"""
import os
import sqlite3
import csv
import io
import json
import secrets
import hashlib
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, Response, session
import openpyxl

app = Flask(__name__, static_folder='static', static_url_path='')

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'orders.db')


def get_secret_key():
    cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.secret_key')
    if os.path.exists(cfg_path):
        with open(cfg_path, 'r') as f:
            return f.read().strip()
    key = secrets.token_hex(32)
    with open(cfg_path, 'w') as f:
        f.write(key)
    return key


app.secret_key = get_secret_key()
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.permanent_session_lifetime = timedelta(days=30)


# ============================================================
# DATABASE
# ============================================================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS orders (
            sub_order_id TEXT PRIMARY KEY,
            platform TEXT NOT NULL,
            order_date TEXT,
            product TEXT,
            sku TEXT,
            catalog_id TEXT,
            size TEXT,
            customer_state TEXT,
            quantity INTEGER DEFAULT 1,
            listed_price REAL DEFAULT 0,
            discounted_price REAL DEFAULT 0,
            status TEXT,
            raw_data TEXT,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS payments (
            sub_order_id TEXT PRIMARY KEY,
            platform TEXT NOT NULL,
            order_date TEXT,
            dispatch_date TEXT,
            payment_date TEXT,
            product TEXT,
            sku TEXT,
            quantity INTEGER DEFAULT 1,
            status TEXT,
            settlement REAL DEFAULT 0,
            sale_amount REAL DEFAULT 0,
            return_amount REAL DEFAULT 0,
            commission REAL DEFAULT 0,
            fixed_fee REAL DEFAULT 0,
            collection_fee REAL DEFAULT 0,
            shipping_fee REAL DEFAULT 0,
            return_shipping REAL DEFAULT 0,
            warehousing_fee REAL DEFAULT 0,
            tcs REAL DEFAULT 0,
            tds REAL DEFAULT 0,
            gst_on_fees REAL DEFAULT 0,
            compensation REAL DEFAULT 0,
            claims REAL DEFAULT 0,
            recovery REAL DEFAULT 0,
            other_fees REAL DEFAULT 0,
            raw_data TEXT,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            salt TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'viewer',
            display_name TEXT,
            created_at TEXT
        );

        CREATE TABLE IF NOT EXISTS api_credentials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT UNIQUE NOT NULL,
            client_id TEXT,
            client_secret TEXT,
            access_token TEXT,
            token_expires_at TEXT,
            extra_config TEXT,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS sync_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT NOT NULL,
            kind TEXT NOT NULL,
            status TEXT NOT NULL,
            message TEXT,
            records_added INTEGER DEFAULT 0,
            records_updated INTEGER DEFAULT 0,
            started_at TEXT,
            finished_at TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_orders_date ON orders(order_date);
        CREATE INDEX IF NOT EXISTS idx_orders_platform ON orders(platform);
        CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status);
        CREATE INDEX IF NOT EXISTS idx_payments_date ON payments(payment_date);
    ''')
    conn.commit()

    for tbl in ('orders', 'payments'):
        try:
            c.execute(f"ALTER TABLE {tbl} ADD COLUMN raw_data TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass

    c.execute("SELECT COUNT(*) as n FROM users")
    if c.fetchone()['n'] == 0:
        create_user('admin', 'admin', role='admin', display_name='Administrator')
        print("=" * 60)
        print("First-time setup: created admin user")
        print("  Username: admin")
        print("  Password: admin")
        print("  CHANGE THIS PASSWORD IMMEDIATELY after logging in.")
        print("=" * 60)

    conn.close()


# ============================================================
# AUTH
# ============================================================
def hash_password(password, salt):
    return hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 200_000).hex()


def create_user(username, password, role='viewer', display_name=None):
    salt = secrets.token_hex(16)
    pw_hash = hash_password(password, salt)
    conn = get_db()
    try:
        conn.execute('''
            INSERT INTO users (username, password_hash, salt, role, display_name, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (username, pw_hash, salt, role, display_name or username, datetime.now().isoformat(timespec='seconds')))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()


def verify_login(username, password):
    conn = get_db()
    row = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    if not row:
        return None
    if hash_password(password, row['salt']) == row['password_hash']:
        return dict(row)
    return None


def change_password(user_id, new_password):
    salt = secrets.token_hex(16)
    pw_hash = hash_password(new_password, salt)
    conn = get_db()
    conn.execute("UPDATE users SET password_hash = ?, salt = ? WHERE id = ?", (pw_hash, salt, user_id))
    conn.commit()
    conn.close()


def current_user():
    if 'user_id' not in session:
        return None
    conn = get_db()
    row = conn.execute("SELECT id, username, role, display_name FROM users WHERE id = ?",
                       (session['user_id'],)).fetchone()
    conn.close()
    return dict(row) if row else None


def login_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if not current_user():
            return jsonify({'error': 'Authentication required'}), 401
        return f(*a, **kw)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        u = current_user()
        if not u:
            return jsonify({'error': 'Authentication required'}), 401
        if u['role'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        return f(*a, **kw)
    return wrapper


# ============================================================
# PARSING HELPERS
# ============================================================
def to_num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v) if not (isinstance(v, float) and (v != v)) else 0.0
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'none', 'null', '-'):
        return 0.0
    s = s.replace(',', '').replace('₹', '')
    try:
        return float(s)
    except (ValueError, InvalidOperation):
        return 0.0


def to_int(v, default=1):
    try:
        return int(to_num(v))
    except (ValueError, TypeError):
        return default


def to_date(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'none', 'null'):
        return ''
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            base = s.split(' ')[0] if ' ' in s and fmt == '%Y-%m-%d' else s
            return datetime.strptime(base, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
        return s[:10]
    return s[:10] if len(s) >= 10 else s


def normalize_status(s):
    if not s:
        return 'Pending'
    t = str(s).lower().strip()
    if 'deliver' in t:
        return 'Delivered'
    if 'return' in t:
        return 'Returned'
    if 'cancel' in t:
        return 'Cancelled'
    if 'exchange' in t:
        return 'Exchange'
    if 'ready' in t or t == 'rts':
        return 'Ready to ship'
    if 'ship' in t or 'dispatch' in t:
        return 'Shipped'
    if 'pend' in t or 'process' in t or 'approve' in t:
        return 'Pending'
    return str(s).strip().title()


def clean_quoted_string(s):
    if not isinstance(s, str):
        return s
    s = s.strip()
    while s.startswith('"') and s.endswith('"') and len(s) >= 2:
        s = s[1:-1]
    return s.strip()


# ============================================================
# PARSERS
# ============================================================
def detect_csv_format(headers):
    h = '|'.join(str(x).lower() for x in headers if x)
    if 'sub order no' in h and 'reason for credit entry' in h:
        return 'meesho_orders'
    if 'order_item_id' in h or 'fsn' in h:
        return 'flipkart_orders'
    if 'order_id' in h and 'platform' in h:
        return 'generic'
    return 'unknown'


def parse_meesho_orders_csv(text):
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for r in reader:
        r = {k.lower().strip(): (v.strip() if v else '') for k, v in r.items() if k}
        sub_id = r.get('sub order no', '').strip()
        if not sub_id:
            continue
        rows.append({
            'sub_order_id': sub_id,
            'platform': 'Meesho',
            'order_date': to_date(r.get('order date')),
            'product': r.get('product name', 'Item'),
            'sku': r.get('sku', ''),
            'catalog_id': r.get('catalog id', ''),
            'size': r.get('size', ''),
            'customer_state': r.get('customer state', ''),
            'quantity': to_int(r.get('quantity'), 1),
            'listed_price': to_num(r.get('supplier listed price (incl. gst + commission)')),
            'discounted_price': to_num(r.get('supplier discounted price (incl gst and commision)')) or to_num(r.get('supplier listed price (incl. gst + commission)')),
            'status': normalize_status(r.get('reason for credit entry')),
            'raw_data': json.dumps(r),
        })
    return rows


def parse_meesho_payment_xlsx(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if 'Order Payments' not in wb.sheetnames:
        raise ValueError(f"'Order Payments' sheet not found. Sheets: {wb.sheetnames}")
    ws = wb['Order Payments']
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return []
    header_row = all_rows[1]
    headers = [str(h).strip() if h else '' for h in header_row]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return -1

    idx = {
        'sub_order': col('Sub Order No'),
        'order_date': col('Order Date'),
        'dispatch_date': col('Dispatch Date'),
        'payment_date': col('Payment Date'),
        'product': col('Product Name'),
        'sku': col('Supplier SKU'),
        'status': col('Live Order Status'),
        'quantity': col('Quantity'),
        'settlement': col('Final Settlement Amount'),
        'sale_amount': col('Total Sale Amount (Incl. Shipping & GST)'),
        'return_amount': col('Total Sale Return Amount (Incl. Shipping & GST)'),
        'commission': col('Meesho Commission (Incl. GST)'),
        'fixed_fee': col('Fixed Fee (Incl. GST)'),
        'warehousing': col('Warehousing fee (Incl. GST)'),
        'shipping': col('Shipping Charge (Incl. GST)'),
        'return_shipping': col('Return Shipping Charge (Incl. GST)'),
        'tcs': col('TCS'),
        'tds': col('TDS'),
        'gst_fees': col('GST on Net Other Support Service Charges'),
        'compensation': col('Compensation'),
        'claims': col('Claims'),
        'recovery': col('Recovery'),
        'other_fees': col('Other Support Service Charges (Excl. GST)'),
    }

    payments = []
    for row in all_rows[2:]:
        if not row:
            continue
        sub_id = row[idx['sub_order']] if 0 <= idx['sub_order'] < len(row) else None
        if sub_id is None:
            continue
        sub_id = str(sub_id).strip()
        if not sub_id or sub_id.lower() in ('nan', 'none'):
            continue
        if '+' in sub_id or 'formula' in sub_id.lower():
            continue
        if 'no data' in sub_id.lower():
            continue

        def get(key):
            i = idx[key]
            return row[i] if 0 <= i < len(row) else None

        raw = {}
        for h_idx, h in enumerate(headers):
            if h and h_idx < len(row) and row[h_idx] is not None:
                v = row[h_idx]
                raw[h] = str(v) if not isinstance(v, (int, float)) else v

        payments.append({
            'sub_order_id': sub_id,
            'platform': 'Meesho',
            'order_date': to_date(get('order_date')),
            'dispatch_date': to_date(get('dispatch_date')),
            'payment_date': to_date(get('payment_date')),
            'product': str(get('product') or ''),
            'sku': str(get('sku') or ''),
            'status': normalize_status(get('status')),
            'quantity': to_int(get('quantity'), 1),
            'settlement': to_num(get('settlement')),
            'sale_amount': to_num(get('sale_amount')),
            'return_amount': to_num(get('return_amount')),
            'commission': to_num(get('commission')),
            'fixed_fee': to_num(get('fixed_fee')),
            'collection_fee': 0,
            'shipping_fee': to_num(get('shipping')),
            'return_shipping': to_num(get('return_shipping')),
            'warehousing_fee': to_num(get('warehousing')),
            'tcs': to_num(get('tcs')),
            'tds': to_num(get('tds')),
            'gst_on_fees': to_num(get('gst_fees')),
            'compensation': to_num(get('compensation')),
            'claims': to_num(get('claims')),
            'recovery': to_num(get('recovery')),
            'other_fees': to_num(get('other_fees')),
            'raw_data': json.dumps(raw, default=str),
        })
    return payments


def parse_flipkart_orders_xlsx(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if 'Orders' not in wb.sheetnames:
        raise ValueError(f"'Orders' sheet not found. Sheets: {wb.sheetnames}")
    ws = wb['Orders']
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    header_row = all_rows[0]
    headers = [str(h).strip().lower() if h else '' for h in header_row]

    def col(name):
        try:
            return headers.index(name.lower())
        except ValueError:
            return -1

    idx = {
        'order_item_id': col('order_item_id'),
        'order_id': col('order_id'),
        'order_date': col('order_date'),
        'status': col('order_item_status'),
        'sku': col('sku'),
        'product': col('product_title'),
        'quantity': col('quantity'),
    }

    orders = []
    for row in all_rows[1:]:
        if not row:
            continue
        sub_id_raw = row[idx['order_item_id']] if 0 <= idx['order_item_id'] < len(row) else None
        if sub_id_raw is None:
            continue
        sub_id = clean_quoted_string(str(sub_id_raw)).strip()
        if not sub_id:
            continue

        def get(key):
            i = idx[key]
            return row[i] if 0 <= i < len(row) else None

        raw = {}
        for h_idx, h in enumerate(headers):
            if h and h_idx < len(row) and row[h_idx] is not None:
                v = row[h_idx]
                raw[h] = clean_quoted_string(str(v)) if isinstance(v, str) else v

        orders.append({
            'sub_order_id': sub_id,
            'platform': 'Flipkart',
            'order_date': to_date(get('order_date')),
            'product': clean_quoted_string(str(get('product') or 'Item')),
            'sku': clean_quoted_string(str(get('sku') or '')),
            'catalog_id': '',
            'size': '',
            'customer_state': '',
            'quantity': to_int(get('quantity'), 1),
            'listed_price': 0,
            'discounted_price': 0,
            'status': normalize_status(get('status')),
            'raw_data': json.dumps(raw, default=str),
        })
    return orders


def parse_flipkart_settlement_xlsx(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if 'Orders' not in wb.sheetnames:
        raise ValueError(f"'Orders' sheet not found. Sheets: {wb.sheetnames}")
    ws = wb['Orders']
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 4:
        return []
    header_row = all_rows[1]
    headers = [str(h).strip() if h else '' for h in header_row]

    def col(*names):
        for name in names:
            for i, h in enumerate(headers):
                if h.lower().startswith(name.lower()):
                    return i
        return -1

    idx = {
        'order_id': col('Order ID'),
        'order_item_id': col('Order item ID'),
        'sale_amount': col('Sale Amount'),
        'commission': col('Commission ('),
        'fixed_fee': col('Fixed Fee'),
        'collection_fee': col('Collection Fee'),
        'shipping_fee': col('Shipping Fee'),
        'reverse_shipping': col('Reverse Shipping'),
        'tcs': col('TCS'),
        'tds': col('TDS'),
        'gst_on_mp': col('GST on MP'),
        'bank_settlement': col('Bank Settlement Value'),
        'order_date': col('Order Date'),
        'dispatch_date': col('Dispatch Date'),
        'payment_date': col('Payment Date'),
        'product_sub_cat': col('Product Sub Category'),
        'sku': col('Seller SKU'),
        'quantity': col('Quantity'),
        'item_return_status': col('Item Return Status'),
        'refund': col('Refund'),
    }

    payments = []
    for row in all_rows[3:]:
        if not row:
            continue
        order_item_id = row[idx['order_item_id']] if 0 <= idx['order_item_id'] < len(row) else None
        if not order_item_id:
            continue
        sub_id = str(order_item_id).strip()
        if not sub_id or sub_id.lower() in ('nan', 'none'):
            continue

        def get(key):
            i = idx[key]
            return row[i] if 0 <= i < len(row) else None

        refund_val = to_num(get('refund'))
        return_status = str(get('item_return_status') or '').strip()
        if return_status and 'return' in return_status.lower():
            status = 'Returned'
        elif refund_val < 0:
            status = 'Returned'
        else:
            status = 'Delivered'

        raw = {}
        for h_idx, h in enumerate(headers):
            if h and h_idx < len(row) and row[h_idx] is not None:
                v = row[h_idx]
                raw[h] = str(v) if not isinstance(v, (int, float)) else v

        payments.append({
            'sub_order_id': sub_id,
            'platform': 'Flipkart',
            'order_date': to_date(get('order_date')),
            'dispatch_date': to_date(get('dispatch_date')),
            'payment_date': to_date(get('payment_date')),
            'product': str(get('product_sub_cat') or ''),
            'sku': str(get('sku') or ''),
            'status': status,
            'quantity': to_int(get('quantity'), 1),
            'settlement': to_num(get('bank_settlement')),
            'sale_amount': to_num(get('sale_amount')),
            'return_amount': refund_val if refund_val < 0 else 0,
            'commission': to_num(get('commission')),
            'fixed_fee': to_num(get('fixed_fee')),
            'collection_fee': to_num(get('collection_fee')),
            'shipping_fee': to_num(get('shipping_fee')),
            'return_shipping': to_num(get('reverse_shipping')),
            'warehousing_fee': 0,
            'tcs': to_num(get('tcs')),
            'tds': to_num(get('tds')),
            'gst_on_fees': to_num(get('gst_on_mp')),
            'compensation': 0,
            'claims': 0,
            'recovery': 0,
            'other_fees': 0,
            'raw_data': json.dumps(raw, default=str),
        })
    return payments


# ============================================================
# DATA OPERATIONS
# ============================================================
def upsert_orders(rows):
    if not rows:
        return {'added': 0, 'updated': 0}
    conn = get_db()
    c = conn.cursor()
    added = updated = 0
    now = datetime.now().isoformat(timespec='seconds')
    for r in rows:
        c.execute("SELECT sub_order_id FROM orders WHERE sub_order_id = ?", (r['sub_order_id'],))
        exists = c.fetchone() is not None
        c.execute('''
            INSERT INTO orders (sub_order_id, platform, order_date, product, sku, catalog_id, size,
                                customer_state, quantity, listed_price, discounted_price, status, raw_data, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sub_order_id) DO UPDATE SET
                platform = excluded.platform,
                order_date = excluded.order_date,
                product = excluded.product,
                sku = excluded.sku,
                catalog_id = excluded.catalog_id,
                size = excluded.size,
                customer_state = excluded.customer_state,
                quantity = excluded.quantity,
                listed_price = excluded.listed_price,
                discounted_price = excluded.discounted_price,
                status = excluded.status,
                raw_data = excluded.raw_data,
                updated_at = excluded.updated_at
        ''', (r['sub_order_id'], r['platform'], r['order_date'], r['product'], r['sku'],
              r.get('catalog_id', ''), r.get('size', ''), r.get('customer_state', ''),
              r['quantity'], r['listed_price'], r['discounted_price'], r['status'],
              r.get('raw_data', ''), now))
        if exists:
            updated += 1
        else:
            added += 1
    conn.commit()
    conn.close()
    return {'added': added, 'updated': updated}


def upsert_payments(rows):
    if not rows:
        return {'added': 0, 'updated': 0, 'synthesized_orders': 0}
    conn = get_db()
    c = conn.cursor()
    added = updated = synthesized = 0
    now = datetime.now().isoformat(timespec='seconds')
    for r in rows:
        c.execute("SELECT sub_order_id FROM payments WHERE sub_order_id = ?", (r['sub_order_id'],))
        exists = c.fetchone() is not None
        c.execute('''
            INSERT INTO payments (sub_order_id, platform, order_date, dispatch_date, payment_date,
                                  product, sku, quantity, status, settlement, sale_amount, return_amount,
                                  commission, fixed_fee, collection_fee, shipping_fee, return_shipping,
                                  warehousing_fee, tcs, tds, gst_on_fees, compensation, claims, recovery,
                                  other_fees, raw_data, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sub_order_id) DO UPDATE SET
                platform = excluded.platform,
                order_date = excluded.order_date,
                dispatch_date = excluded.dispatch_date,
                payment_date = excluded.payment_date,
                product = excluded.product,
                sku = excluded.sku,
                quantity = excluded.quantity,
                status = excluded.status,
                settlement = excluded.settlement,
                sale_amount = excluded.sale_amount,
                return_amount = excluded.return_amount,
                commission = excluded.commission,
                fixed_fee = excluded.fixed_fee,
                collection_fee = excluded.collection_fee,
                shipping_fee = excluded.shipping_fee,
                return_shipping = excluded.return_shipping,
                warehousing_fee = excluded.warehousing_fee,
                tcs = excluded.tcs,
                tds = excluded.tds,
                gst_on_fees = excluded.gst_on_fees,
                compensation = excluded.compensation,
                claims = excluded.claims,
                recovery = excluded.recovery,
                other_fees = excluded.other_fees,
                raw_data = excluded.raw_data,
                updated_at = excluded.updated_at
        ''', (r['sub_order_id'], r['platform'], r['order_date'], r['dispatch_date'], r['payment_date'],
              r['product'], r['sku'], r['quantity'], r['status'], r['settlement'], r['sale_amount'],
              r['return_amount'], r['commission'], r['fixed_fee'], r['collection_fee'], r['shipping_fee'],
              r['return_shipping'], r['warehousing_fee'], r['tcs'], r['tds'], r['gst_on_fees'],
              r['compensation'], r['claims'], r['recovery'], r['other_fees'],
              r.get('raw_data', ''), now))
        if exists:
            updated += 1
        else:
            added += 1

        c.execute("SELECT 1 FROM orders WHERE sub_order_id = ?", (r['sub_order_id'],))
        if not c.fetchone():
            c.execute('''
                INSERT INTO orders (sub_order_id, platform, order_date, product, sku,
                                    quantity, listed_price, discounted_price, status, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (r['sub_order_id'], r['platform'], r['order_date'], r['product'], r['sku'],
                  r['quantity'], 0, 0, r['status'], now))
            synthesized += 1
        else:
            c.execute("UPDATE orders SET status = ? WHERE sub_order_id = ?",
                      (r['status'], r['sub_order_id']))

    conn.commit()
    conn.close()
    return {'added': added, 'updated': updated, 'synthesized_orders': synthesized}


# ============================================================
# CALCULATIONS
# ============================================================
def compute_dashboard(date_from='', date_to='', platform=''):
    conn = get_db()
    c = conn.cursor()
    where = []
    params = []
    if date_from:
        where.append("o.order_date >= ?"); params.append(date_from)
    if date_to:
        where.append("o.order_date <= ?"); params.append(date_to)
    if platform:
        where.append("o.platform = ?"); params.append(platform)
    where_sql = ' WHERE ' + ' AND '.join(where) if where else ''

    c.execute(f'SELECT o.status, COUNT(*) as n FROM orders o {where_sql} GROUP BY o.status', params)
    status_counts = {r['status']: r['n'] for r in c.fetchall()}

    c.execute(f'SELECT COUNT(*) as n FROM orders o {where_sql}', params)
    total = c.fetchone()['n']

    c.execute(f'SELECT o.platform, COUNT(*) as n FROM orders o {where_sql} GROUP BY o.platform', params)
    platform_counts = {r['platform']: r['n'] for r in c.fetchall()}

    c.execute(f'''
        SELECT
            COALESCE(SUM(p.sale_amount), 0) as sale_amount,
            COALESCE(SUM(p.return_amount), 0) as return_amount,
            COALESCE(SUM(p.settlement), 0) as settlement,
            COALESCE(SUM(p.commission), 0) as commission,
            COALESCE(SUM(p.fixed_fee), 0) as fixed_fee,
            COALESCE(SUM(p.collection_fee), 0) as collection_fee,
            COALESCE(SUM(p.shipping_fee), 0) as shipping_fee,
            COALESCE(SUM(p.return_shipping), 0) as return_shipping,
            COALESCE(SUM(p.warehousing_fee), 0) as warehousing_fee,
            COALESCE(SUM(p.tcs), 0) as tcs,
            COALESCE(SUM(p.tds), 0) as tds,
            COALESCE(SUM(p.gst_on_fees), 0) as gst_on_fees,
            COALESCE(SUM(p.compensation), 0) as compensation,
            COALESCE(SUM(p.claims), 0) as claims,
            COALESCE(SUM(p.recovery), 0) as recovery,
            COALESCE(SUM(p.other_fees), 0) as other_fees,
            COUNT(*) as paid_count
        FROM payments p
        JOIN orders o ON o.sub_order_id = p.sub_order_id
        {where_sql}
    ''', params)
    pay = dict(c.fetchone())

    c.execute(f'''
        SELECT o.platform, COUNT(*) as count,
               COALESCE(SUM(p.sale_amount), 0) as sales,
               COALESCE(SUM(p.return_amount), 0) as returns,
               COALESCE(SUM(p.settlement), 0) as settlement
        FROM orders o
        LEFT JOIN payments p ON p.sub_order_id = o.sub_order_id
        {where_sql}
        GROUP BY o.platform
    ''', params)
    platforms = {r['platform']: dict(r) for r in c.fetchall()}

    est_where = list(where)
    est_params = list(params)
    c.execute(f'''
        SELECT COALESCE(SUM(o.discounted_price * o.quantity), 0) as est_gross
        FROM orders o
        LEFT JOIN payments p ON p.sub_order_id = o.sub_order_id
        WHERE p.sub_order_id IS NULL AND o.status = 'Delivered'
        {' AND ' + ' AND '.join(est_where) if est_where else ''}
    ''', est_params)
    est_unpaid_gross = c.fetchone()['est_gross']

    monthly_where = list(where) + ["substr(o.order_date, 1, 7) != ''"]
    monthly_params = list(params)
    c.execute(f'''
        SELECT substr(o.order_date, 1, 7) as month,
               COUNT(*) as orders,
               COALESCE(SUM(p.sale_amount), 0) as sales,
               COALESCE(SUM(p.settlement), 0) as settlement
        FROM orders o
        LEFT JOIN payments p ON p.sub_order_id = o.sub_order_id
        WHERE {' AND '.join(monthly_where)}
        GROUP BY month
        ORDER BY month
    ''', monthly_params)
    monthly = [dict(r) for r in c.fetchall()]

    c.execute(f'''
        SELECT o.product as product, COUNT(*) as count,
               COALESCE(SUM(p.sale_amount), 0) as revenue,
               COALESCE(SUM(p.settlement), 0) as settlement,
               COALESCE(SUM(o.discounted_price * o.quantity), 0) as est_revenue,
               SUM(CASE WHEN o.status = 'Returned' THEN 1 ELSE 0 END) as returns
        FROM orders o
        LEFT JOIN payments p ON p.sub_order_id = o.sub_order_id
        {where_sql}
        GROUP BY o.product
        ORDER BY revenue DESC, est_revenue DESC
        LIMIT 10
    ''', params)
    top_products = []
    for r in c.fetchall():
        d = dict(r)
        d['display_revenue'] = d['revenue'] if d['revenue'] > 0 else d['est_revenue']
        top_products.append(d)

    conn.close()

    delivered = status_counts.get('Delivered', 0)
    returned = status_counts.get('Returned', 0)
    cancelled = status_counts.get('Cancelled', 0)
    pending = status_counts.get('Pending', 0)
    ready = status_counts.get('Ready to ship', 0)
    shipped = status_counts.get('Shipped', 0)
    in_transit = shipped + ready + pending

    gross_sales = pay['sale_amount'] + est_unpaid_gross
    return_amount = abs(pay['return_amount'])
    net_settlement = pay['settlement']
    total_fees = (abs(pay['commission']) + abs(pay['fixed_fee']) + abs(pay['collection_fee']) +
                  abs(pay['shipping_fee']) + abs(pay['return_shipping']) + abs(pay['warehousing_fee']) +
                  abs(pay['gst_on_fees']) + abs(pay['other_fees']))
    total_tax = abs(pay['tcs']) + abs(pay['tds'])

    aov = gross_sales / delivered if delivered > 0 else 0
    avg_settlement = net_settlement / pay['paid_count'] if pay['paid_count'] > 0 else 0
    return_rate = (returned / total * 100) if total > 0 else 0
    net_margin = (net_settlement / gross_sales * 100) if gross_sales > 0 else 0

    return {
        'order_counts': {
            'total': total, 'delivered': delivered, 'in_transit': in_transit,
            'returned': returned, 'cancelled': cancelled,
            'pending': pending, 'ready_to_ship': ready, 'shipped': shipped,
        },
        'platform_counts': platform_counts,
        'money': {
            'gross_sales': round(gross_sales, 2),
            'return_amount': round(return_amount, 2),
            'net_settlement': round(net_settlement, 2),
            'outstanding': round(net_settlement, 2),
            'commission': round(abs(pay['commission']), 2),
            'shipping': round(abs(pay['shipping_fee']), 2),
            'return_shipping': round(abs(pay['return_shipping']), 2),
            'fixed_fee': round(abs(pay['fixed_fee']), 2),
            'collection_fee': round(abs(pay['collection_fee']), 2),
            'warehousing_fee': round(abs(pay['warehousing_fee']), 2),
            'total_fees': round(total_fees, 2),
            'tcs': round(abs(pay['tcs']), 2),
            'tds': round(abs(pay['tds']), 2),
            'total_tax': round(total_tax, 2),
            'compensation': round(pay['compensation'], 2),
            'claims': round(pay['claims'], 2),
            'recovery': round(pay['recovery'], 2),
        },
        'derived': {
            'aov': round(aov, 2),
            'avg_settlement': round(avg_settlement, 2),
            'return_rate': round(return_rate, 1),
            'net_margin': round(net_margin, 1),
            'paid_count': pay['paid_count'],
        },
        'platform_breakdown': {
            p: {
                'count': platforms.get(p, {}).get('count', 0),
                'sales': round(platforms.get(p, {}).get('sales', 0), 2),
                'returns': round(abs(platforms.get(p, {}).get('returns', 0)), 2),
                'settlement': round(platforms.get(p, {}).get('settlement', 0), 2),
            }
            for p in ('Meesho', 'Flipkart')
        },
        'monthly_trend': monthly,
        'top_products': top_products,
    }


def list_orders(date_from='', date_to='', platform='', status='', search=''):
    conn = get_db()
    c = conn.cursor()
    where = []
    params = []
    if date_from:
        where.append("o.order_date >= ?"); params.append(date_from)
    if date_to:
        where.append("o.order_date <= ?"); params.append(date_to)
    if platform:
        where.append("o.platform = ?"); params.append(platform)
    if status:
        where.append("o.status = ?"); params.append(status)
    if search:
        where.append("(o.product LIKE ? OR o.sub_order_id LIKE ? OR o.sku LIKE ?)")
        params.extend([f'%{search}%'] * 3)
    where_sql = ' WHERE ' + ' AND '.join(where) if where else ''

    c.execute(f'''
        SELECT o.sub_order_id, o.platform, o.order_date, o.product, o.sku, o.catalog_id,
               o.size, o.customer_state, o.quantity, o.listed_price, o.discounted_price,
               o.status, o.updated_at,
               p.settlement, p.sale_amount, p.commission, p.payment_date, p.dispatch_date
        FROM orders o
        LEFT JOIN payments p ON p.sub_order_id = o.sub_order_id
        {where_sql}
        ORDER BY o.order_date DESC
    ''', params)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def get_order_detail(sub_order_id):
    conn = get_db()
    c = conn.cursor()
    order = c.execute("SELECT * FROM orders WHERE sub_order_id = ?", (sub_order_id,)).fetchone()
    if not order:
        conn.close()
        return None
    payment = c.execute("SELECT * FROM payments WHERE sub_order_id = ?", (sub_order_id,)).fetchone()
    conn.close()
    result = {'order': dict(order), 'payment': dict(payment) if payment else None}
    if result['order'].get('raw_data'):
        try:
            result['order']['raw_data_parsed'] = json.loads(result['order']['raw_data'])
        except Exception:
            pass
    if result['payment'] and result['payment'].get('raw_data'):
        try:
            result['payment']['raw_data_parsed'] = json.loads(result['payment']['raw_data'])
        except Exception:
            pass
    return result


def list_payments(search=''):
    conn = get_db()
    c = conn.cursor()
    if search:
        c.execute('''SELECT * FROM payments WHERE sub_order_id LIKE ? OR product LIKE ? OR status LIKE ?
                     ORDER BY payment_date DESC''', [f'%{search}%'] * 3)
    else:
        c.execute('SELECT * FROM payments ORDER BY payment_date DESC')
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


# ============================================================
# API INTEGRATION
# ============================================================
def get_api_creds(platform):
    conn = get_db()
    row = conn.execute("SELECT * FROM api_credentials WHERE platform = ?", (platform,)).fetchone()
    conn.close()
    return dict(row) if row else None


def save_api_creds(platform, client_id=None, client_secret=None, access_token=None):
    conn = get_db()
    now = datetime.now().isoformat(timespec='seconds')
    conn.execute('''
        INSERT INTO api_credentials (platform, client_id, client_secret, access_token, updated_at)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(platform) DO UPDATE SET
            client_id = COALESCE(excluded.client_id, api_credentials.client_id),
            client_secret = COALESCE(excluded.client_secret, api_credentials.client_secret),
            access_token = COALESCE(excluded.access_token, api_credentials.access_token),
            updated_at = excluded.updated_at
    ''', (platform, client_id, client_secret, access_token, now))
    conn.commit()
    conn.close()


def log_sync(platform, kind, status, message='', added=0, updated=0):
    conn = get_db()
    now = datetime.now().isoformat(timespec='seconds')
    conn.execute('''INSERT INTO sync_log (platform, kind, status, message, records_added, records_updated, started_at, finished_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                 (platform, kind, status, message, added, updated, now, now))
    conn.commit()
    conn.close()


def fetch_flipkart_orders_via_api():
    """Calls the official Flipkart Marketplace API to fetch pending orders.
    Requires Client ID and Client Secret obtained from Flipkart Seller Hub → API Access.
    Returns a list of order dicts ready for upsert_orders()."""
    import urllib.request
    import base64

    creds = get_api_creds('flipkart')
    if not creds or not creds.get('client_id') or not creds.get('client_secret'):
        raise ValueError('Flipkart API credentials not configured. Add them under API Setup.')

    auth_b64 = base64.b64encode(f"{creds['client_id']}:{creds['client_secret']}".encode()).decode()
    token_url = 'https://api.flipkart.net/oauth-service/oauth/token?grant_type=client_credentials&scope=Seller_Api'
    req = urllib.request.Request(token_url, method='POST')
    req.add_header('Authorization', f'Basic {auth_b64}')
    with urllib.request.urlopen(req, timeout=30) as resp:
        token_data = json.loads(resp.read().decode())
    access_token = token_data['access_token']

    search_url = 'https://api.flipkart.net/sellers/v3/shipments/filter'
    payload = json.dumps({
        'filter': {'states': ['APPROVED', 'PACKED']},
        'pagination': {'pageSize': 50},
    }).encode()
    req = urllib.request.Request(search_url, data=payload, method='POST')
    req.add_header('Authorization', f'Bearer {access_token}')
    req.add_header('Content-Type', 'application/json')
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read().decode())

    shipments = data.get('shipments', [])
    orders = []
    for s in shipments:
        item_id = s.get('orderItemId') or (s.get('orderItem') or {}).get('orderItemId')
        if not item_id:
            continue
        oi = (s.get('orderItems') or [{}])[0] if s.get('orderItems') else {}
        state = oi.get('status') or 'Pending'
        orders.append({
            'sub_order_id': str(item_id),
            'platform': 'Flipkart',
            'order_date': to_date(s.get('orderDate')),
            'product': oi.get('title', 'Item'),
            'sku': oi.get('sku', ''),
            'catalog_id': '',
            'size': '',
            'customer_state': '',
            'quantity': to_int(oi.get('quantity'), 1),
            'listed_price': to_num(oi.get('sellingPrice')),
            'discounted_price': to_num(oi.get('sellingPrice')),
            'status': normalize_status(state),
            'raw_data': json.dumps(s, default=str),
        })
    return orders


def fetch_meesho_orders_via_api():
    """Calls the Meesho Supplier API. Requires API token obtained from supplier-api@meesho.com."""
    import urllib.request

    creds = get_api_creds('meesho')
    if not creds or not creds.get('access_token'):
        raise ValueError('Meesho API token not configured. Add it under API Setup.')

    url = 'https://api.meesho.com/supplier/v1/orders?status=PENDING&limit=50'
    req = urllib.request.Request(url)
    req.add_header('Authorization', f'Bearer {creds["access_token"]}')
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read().decode())

    raw_orders = data.get('orders') or data.get('data') or []
    orders = []
    for o in raw_orders:
        sub_id = o.get('sub_order_id') or o.get('order_id')
        if not sub_id:
            continue
        orders.append({
            'sub_order_id': str(sub_id),
            'platform': 'Meesho',
            'order_date': to_date(o.get('order_date')),
            'product': o.get('product_name', 'Item'),
            'sku': o.get('sku', ''),
            'catalog_id': str(o.get('catalog_id', '')),
            'size': o.get('size', ''),
            'customer_state': o.get('customer_state', ''),
            'quantity': to_int(o.get('quantity'), 1),
            'listed_price': to_num(o.get('listing_price')),
            'discounted_price': to_num(o.get('transfer_price') or o.get('selling_price')),
            'status': normalize_status(o.get('status')),
            'raw_data': json.dumps(o, default=str),
        })
    return orders


# ============================================================
# ROUTES
# ============================================================
@app.route('/')
def index():
    if not current_user():
        return send_from_directory('static', 'login.html')
    return send_from_directory('static', 'index.html')


@app.route('/login.html')
def login_page():
    return send_from_directory('static', 'login.html')


# ---------- Auth ----------
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    user = verify_login(data.get('username', ''), data.get('password', ''))
    if not user:
        return jsonify({'error': 'Invalid username or password'}), 401
    session.permanent = True
    session['user_id'] = user['id']
    return jsonify({
        'id': user['id'],
        'username': user['username'],
        'role': user['role'],
        'display_name': user['display_name'],
        'must_change_password': user['username'] == 'admin' and verify_login('admin', 'admin') is not None,
    })


@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'ok': True})


@app.route('/api/auth/me')
def api_me():
    u = current_user()
    if not u:
        return jsonify({'error': 'Not logged in'}), 401
    return jsonify(u)


@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.get_json() or {}
    u = current_user()
    if not verify_login(u['username'], data.get('current_password', '')):
        return jsonify({'error': 'Current password is incorrect'}), 400
    new_pw = data.get('new_password', '')
    if len(new_pw) < 6:
        return jsonify({'error': 'New password must be at least 6 characters'}), 400
    change_password(u['id'], new_pw)
    return jsonify({'ok': True})


@app.route('/api/users', methods=['GET'])
@admin_required
def api_list_users():
    conn = get_db()
    rows = conn.execute("SELECT id, username, role, display_name, created_at FROM users ORDER BY id").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    role = data.get('role', 'viewer')
    if not username or len(password) < 6:
        return jsonify({'error': 'Username required and password must be at least 6 characters'}), 400
    if role not in ('admin', 'viewer'):
        return jsonify({'error': 'Invalid role'}), 400
    if not create_user(username, password, role=role, display_name=data.get('display_name') or username):
        return jsonify({'error': 'Username already exists'}), 400
    return jsonify({'ok': True})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_delete_user(user_id):
    me = current_user()
    if me['id'] == user_id:
        return jsonify({'error': 'Cannot delete yourself'}), 400
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ---------- Data ----------
@app.route('/api/upload', methods=['POST'])
@admin_required
def api_upload():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file uploaded'}), 400
    filename = file.filename
    fn_low = filename.lower()
    data = file.read()
    try:
        if fn_low.endswith('.csv'):
            text = data.decode('utf-8-sig')
            first_line = text.split('\n', 1)[0]
            headers = [h.strip().strip('"').lower() for h in first_line.split(',')]
            fmt = detect_csv_format(headers)
            if fmt == 'meesho_orders':
                rows = parse_meesho_orders_csv(text)
                result = upsert_orders(rows)
                return jsonify({'file': filename, 'kind': 'Meesho Orders', 'records': len(rows), **result})
            return jsonify({'error': f'Unrecognized CSV format. Detected headers: {headers[:5]}'}), 400

        if fn_low.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            sheets = wb.sheetnames
            wb.close()
            if 'Order Payments' in sheets:
                rows = parse_meesho_payment_xlsx(data)
                result = upsert_payments(rows)
                return jsonify({'file': filename, 'kind': 'Meesho Payment', 'records': len(rows), **result})
            if 'Orders' in sheets and 'Help' in sheets and 'Summary of report' not in sheets:
                rows = parse_flipkart_orders_xlsx(data)
                result = upsert_orders(rows)
                return jsonify({'file': filename, 'kind': 'Flipkart Orders', 'records': len(rows), **result})
            if 'Summary of report' in sheets:
                rows = parse_flipkart_settlement_xlsx(data)
                result = upsert_payments(rows)
                return jsonify({'file': filename, 'kind': 'Flipkart Settlement', 'records': len(rows), **result})
            return jsonify({'error': f'Unrecognized XLSX format. Sheets: {sheets}'}), 400

        return jsonify({'error': 'Only .csv, .xlsx, and .xls files supported'}), 400
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/dashboard')
@login_required
def api_dashboard():
    return jsonify(compute_dashboard(
        date_from=request.args.get('from', ''),
        date_to=request.args.get('to', ''),
        platform=request.args.get('platform', ''),
    ))


@app.route('/api/orders')
@login_required
def api_orders():
    return jsonify(list_orders(
        date_from=request.args.get('from', ''),
        date_to=request.args.get('to', ''),
        platform=request.args.get('platform', ''),
        status=request.args.get('status', ''),
        search=request.args.get('search', ''),
    ))


@app.route('/api/orders/<path:sub_order_id>')
@login_required
def api_order_detail(sub_order_id):
    result = get_order_detail(sub_order_id)
    if not result:
        return jsonify({'error': 'Order not found'}), 404
    return jsonify(result)


@app.route('/api/payments')
@login_required
def api_payments():
    return jsonify(list_payments(search=request.args.get('search', '')))


@app.route('/api/orders/<path:sub_order_id>', methods=['DELETE'])
@admin_required
def api_delete_order(sub_order_id):
    conn = get_db()
    conn.execute("DELETE FROM orders WHERE sub_order_id = ?", (sub_order_id,))
    conn.execute("DELETE FROM payments WHERE sub_order_id = ?", (sub_order_id,))
    conn.commit()
    conn.close()
    return jsonify({'deleted': sub_order_id})


@app.route('/api/clear', methods=['POST'])
@admin_required
def api_clear():
    conn = get_db()
    conn.execute("DELETE FROM orders")
    conn.execute("DELETE FROM payments")
    conn.commit()
    conn.close()
    return jsonify({'cleared': True})


@app.route('/api/export.csv')
@login_required
def api_export():
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        SELECT o.sub_order_id, o.order_date, o.platform, o.product, o.sku, o.quantity,
               o.discounted_price, o.status, o.customer_state,
               p.settlement, p.sale_amount, p.return_amount, p.commission,
               p.shipping_fee, p.return_shipping, p.tcs, p.tds, p.payment_date
        FROM orders o
        LEFT JOIN payments p ON p.sub_order_id = o.sub_order_id
        ORDER BY o.order_date DESC
    ''')
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['sub_order_id', 'order_date', 'platform', 'product', 'sku', 'quantity',
                     'price', 'status', 'customer_state', 'settlement', 'sale_amount',
                     'return_amount', 'commission', 'shipping_fee', 'return_shipping',
                     'tcs', 'tds', 'payment_date'])
    for r in c.fetchall():
        writer.writerow([r[k] if r[k] is not None else '' for k in r.keys()])
    conn.close()
    return Response(output.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename=orders_export_{datetime.now().strftime("%Y%m%d")}.csv'})


# ---------- API integration ----------
@app.route('/api/api-credentials', methods=['GET'])
@admin_required
def api_get_credentials():
    conn = get_db()
    rows = conn.execute("SELECT platform, client_id, access_token, updated_at FROM api_credentials").fetchall()
    conn.close()
    out = {}
    for r in rows:
        out[r['platform']] = {
            'has_client_id': bool(r['client_id']),
            'has_client_secret': bool(r['client_id']),
            'has_access_token': bool(r['access_token']),
            'updated_at': r['updated_at'],
        }
    return jsonify(out)


@app.route('/api/api-credentials/<platform>', methods=['POST'])
@admin_required
def api_save_credentials(platform):
    if platform not in ('meesho', 'flipkart'):
        return jsonify({'error': 'Unknown platform'}), 400
    data = request.get_json() or {}
    save_api_creds(
        platform,
        client_id=data.get('client_id') or None,
        client_secret=data.get('client_secret') or None,
        access_token=data.get('access_token') or None,
    )
    return jsonify({'ok': True})


@app.route('/api/sync/<platform>', methods=['POST'])
@admin_required
def api_sync_platform(platform):
    try:
        if platform == 'flipkart':
            orders = fetch_flipkart_orders_via_api()
        elif platform == 'meesho':
            orders = fetch_meesho_orders_via_api()
        else:
            return jsonify({'error': 'Unknown platform'}), 400
        result = upsert_orders(orders)
        log_sync(platform, 'orders_pull', 'success', f'{len(orders)} orders',
                 result['added'], result['updated'])
        return jsonify({'fetched': len(orders), **result})
    except Exception as e:
        log_sync(platform, 'orders_pull', 'error', str(e))
        return jsonify({'error': str(e)}), 500


@app.route('/api/sync-log')
@login_required
def api_sync_log():
    conn = get_db()
    rows = conn.execute("SELECT * FROM sync_log ORDER BY id DESC LIMIT 50").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


if __name__ == '__main__':
    init_db()
    print('=' * 60)
    print('Order Tracker — running at http://localhost:5000')
    print('Default login: admin / admin (change immediately)')
    print('=' * 60)
    app.run(host='127.0.0.1', port=5000, debug=False)
